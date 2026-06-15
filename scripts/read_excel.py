import openpyxl
import os
import json

file_path = 'report6.xls'

# Check if file exists
if not os.path.exists(file_path):
    print(f"File not found: {file_path}")
    exit(1)

try:
    # Try loading as XLSX first (sometimes .xls is actually .xlsx)
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    print(f"✅ Excel loaded successfully")
    print(f"Sheet name: {ws.title}")

    # Extract first 15 rows
    print("\n=== FIRST 15 ROWS ===\n")
    rows_data = []
    for i in range(1, min(16, ws.max_row + 1)):
        row_values = []
        for cell in ws[i]:
            row_values.append(
                str(cell.value) if cell.value is not None else "-")
        rows_data.append(row_values)
        print(f"Row {i}: {row_values}")

    # Get max row
    print(f"\n=== FILE STATS ===")
    print(f"Total rows: {ws.max_row}")
    print(f"Total columns: {ws.max_column}")

    # Last row
    if ws.max_row > 15:
        print(f"\nLast row (#{ws.max_row}):")
        last_row_values = []
        for cell in ws[ws.max_row]:
            val = str(cell.value) if cell.value is not None else "-"
            last_row_values.append(val)
            print(f"  {val}")

except Exception as e:
    print(f"❌ Error: {type(e).__name__}: {e}")
    print(f"File might be in old .xls format, not .xlsx")

    # Try alternative approach with xlrd
    try:
        import xlrd
        book = xlrd.open_workbook('report6.xls')
        sheet = book.sheet_by_index(0)

        print(f"\n✅ Loaded with xlrd")
        print(f"Sheet: {sheet.name}")
        print(f"Rows: {sheet.nrows}, Cols: {sheet.ncols}")

        for i in range(min(15, sheet.nrows)):
            print(f"Row {i+1}: {sheet.row_values(i)}")

    except ImportError:
        print("xlrd not available, installing...")
